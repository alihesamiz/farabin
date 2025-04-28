from openpyxl import load_workbook
from celery import shared_task
import logging
import cohere

from django.db.transaction import atomic
from django.conf import settings

from management.models import HumanResource, PersonelInformation, SWOTMatrix, SWOTAnalysis


logger = logging.getLogger("management")


@shared_task(bind=True,)
def process_personnel_excel(self, id: int):
    try:
        hr_instance = HumanResource.objects.get(id=id)
        file_path = hr_instance.excel_file.path
        company_name = hr_instance.company.company_title

        logger.info(f"Processing Excel file for company: {company_name}")
        logger.info(f"Loading workbook from {file_path}")

        wb = load_workbook(file_path)
        sheet = wb.active

        logger.info(f"Loading rows of data from {company_name} excel")

        personnel_list = []
        position_map = {}  # Maps position to list of PersonelInformation
        # Stores (person, reports_to_positions, cooperates_with_positions)
        person_relations = []

        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not any(row):
                continue

            name, position, reports_to_position, cooperates_with_position, obligations = row[
                :5]

            if not (name and obligations and position):
                logger.warning(
                    f"Skipped invalid record: {name}, {obligations}, {position}. "
                    f"Missing reports_to {reports_to_position} or cooperates_with {cooperates_with_position}"
                )
                continue

            # Normalize data
            name = name.strip().title()
            position = position.strip().upper()
            obligations = obligations.strip()
            reports_to_positions = (
                [p.strip().upper()
                 for p in reports_to_position.strip().split(",")]
                if reports_to_position
                else []
            )
            cooperates_with_positions = (
                [p.strip().upper()
                 for p in cooperates_with_position.strip().split(",")]
                if cooperates_with_position
                else []
            )

            person = PersonelInformation(
                human_resource=hr_instance,
                name=name,
                obligations=obligations,
                position=position,
            )
            personnel_list.append(person)

            # Store person in position_map (as a list)
            if position not in position_map:
                position_map[position] = []
            position_map[position].append(person)

            # Store relationships for later processing
            person_relations.append(
                (person, reports_to_positions, cooperates_with_positions))

        with atomic():
            # Bulk create personnel
            created_personnel = PersonelInformation.objects.bulk_create(
                personnel_list)

            # Set many-to-many relationships
            for person, reports_to_positions, cooperates_with_positions in person_relations:
                # Set reports_to relationships
                for pos in reports_to_positions:
                    if pos in position_map:
                        person.reports_to.add(*position_map[pos])

                # Set cooperates_with relationships
                for pos in cooperates_with_positions:
                    if pos in position_map:
                        person.cooperates_with.add(*position_map[pos])

        logger.info(
            f"Successfully processed {len(created_personnel)} personnel records")

    except HumanResource.DoesNotExist:
        logger.error(f"HumanResource with id {id} not found")
        raise
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        raise


@shared_task(bind=True, rate_limit='5/m')
def generate_swot_analysis(self, id: int):
    try:
        swot_instance = SWOTMatrix.objects.get(id=id)
        logger.info(f"Creating SWOT analysis for {swot_instance}")

        results = {}
        values = {}

        analysis_groups = {
            "so":("strengths", "opportunities"),
            "st":("strengths", "threats"),
            "wo":("weaknesses", "opportunities"),
            "wt":("weaknesses", "threats"),
        }

        for field in ["strengths", "weaknesses", "opportunities", "threats"]:
            related_qs = getattr(swot_instance, field).values(
                "name", "custom_name")
            values[field] = [
                item["custom_name"] or item["name"]
                for item in related_qs
                if item["custom_name"] or item["name"]
            ]

        grouped_results = {}

        for group1, group2 in analysis_groups:
            key = f"{group1}_{group2}"
            grouped_results[key] = {
                group1: values.get(group1, []),
                group2: values.get(group2, [])
            }
        print(values)
        print(grouped_results)
        # generator = cohere.ClientV2(settings.FARABIN_COHERE_API_KEY)
        # prompt = f"""
        # Provide the strategic and collision points for the SWOT matrix given the strengths, threats, opportunities, and weaknesses presented as below:
        #     Strengths: {results['strengths']}
        #     Weaknesses: {results['weaknesses']}
        #     Opportunities: {results['opportunities']}
        #     Threats: {results['threats']}
        # this report and analysis should be in Persian language and a professional format and should be suitable for a business report.
        # """
        # response = generator.chat(
        #     model="command-r-plus",
        #     messages=[
        #         {
        #             "role": "user",
        #             "content": f"{prompt}"}],
        # ).message.content[0].text

        # SWOTAnalysis.objects.create(
        #     swot_matrix=swot_instance,
        #     analysis=response,
        #     is_approved=False
        # )

        logger.info(f"SWOT analysis created for {swot_instance}")

    except Exception as e:
        logger.error(
            f"Error while creating SWOT analysis for {id}: {str(e)}")
        return str(e)
